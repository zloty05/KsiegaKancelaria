"""Warstwa bazy danych SQLite dla rejestru pism.

Tabela `pisma` przechowuje wszystkie zarejestrowane pisma (przychodzące i wychodzące).
Eksport do Excela przez openpyxl.
"""

from __future__ import annotations

import logging
import sqlite3
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

from config import DB_PATH

logger = logging.getLogger(__name__)

# Kolumny w kolejności wyświetlania/eksportu.
COLUMNS = [
    "id",
    "data_wpisu",
    "data_pisma",
    "sygnatura",
    "sad",
    "typ_pisma",
    "strona_powodowa",
    "strona_pozwana",
    "kierunek",
    "sciezka_pliku",
    "nazwa_pliku",
]

# Polskie nagłówki dla eksportu Excel.
EXCEL_HEADERS = {
    "id": "ID",
    "data_wpisu": "Data wpisu",
    "data_pisma": "Data pisma",
    "sygnatura": "Sygnatura akt",
    "sad": "Sąd",
    "typ_pisma": "Typ pisma",
    "strona_powodowa": "Strona powodowa",
    "strona_pozwana": "Strona pozwana",
    "kierunek": "Kierunek",
    "sciezka_pliku": "Ścieżka pliku",
    "nazwa_pliku": "Nazwa pliku",
}


@dataclass
class Pismo:
    """Rekord pisma z bazy."""

    id: Optional[int]
    data_wpisu: str
    data_pisma: Optional[str]
    sygnatura: Optional[str]
    sad: Optional[str]
    typ_pisma: Optional[str]
    strona_powodowa: Optional[str]
    strona_pozwana: Optional[str]
    kierunek: str  # 'IN' lub 'OUT'
    sciezka_pliku: Optional[str]
    nazwa_pliku: Optional[str]


def _connect(db_path: Path = DB_PATH) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn


def init_db(db_path: Path = DB_PATH) -> None:
    """Tworzy tabelę pisma, jeśli nie istnieje."""
    with _connect(db_path) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS pisma (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                data_wpisu TEXT NOT NULL,
                data_pisma TEXT,
                sygnatura TEXT,
                sad TEXT,
                typ_pisma TEXT,
                strona_powodowa TEXT,
                strona_pozwana TEXT,
                kierunek TEXT NOT NULL CHECK(kierunek IN ('IN', 'OUT')),
                sciezka_pliku TEXT,
                nazwa_pliku TEXT
            )
            """
        )
        conn.commit()
    logger.info("Baza danych zainicjalizowana: %s", db_path)


def add_pismo(
    *,
    data_pisma: Optional[str],
    sygnatura: Optional[str],
    sad: Optional[str],
    typ_pisma: Optional[str],
    strona_powodowa: Optional[str],
    strona_pozwana: Optional[str],
    kierunek: str,
    sciezka_pliku: Optional[str],
    nazwa_pliku: Optional[str],
    db_path: Path = DB_PATH,
) -> int:
    """Dodaje pismo do rejestru. Zwraca id nowego wpisu."""
    if kierunek not in ("IN", "OUT"):
        raise ValueError(f"Nieprawidłowy kierunek: {kierunek!r} (oczekiwano 'IN'/'OUT')")
    data_wpisu = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with _connect(db_path) as conn:
        cur = conn.execute(
            """
            INSERT INTO pisma (
                data_wpisu, data_pisma, sygnatura, sad, typ_pisma,
                strona_powodowa, strona_pozwana, kierunek, sciezka_pliku, nazwa_pliku
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                data_wpisu,
                data_pisma,
                sygnatura,
                sad,
                typ_pisma,
                strona_powodowa,
                strona_pozwana,
                kierunek,
                sciezka_pliku,
                nazwa_pliku,
            ),
        )
        conn.commit()
        return cur.lastrowid


def _row_to_pismo(row: sqlite3.Row) -> Pismo:
    return Pismo(**{k: row[k] for k in COLUMNS})


def get_all_pisma(db_path: Path = DB_PATH) -> list[Pismo]:
    """Zwraca wszystkie pisma, najnowsze pierwsze."""
    with _connect(db_path) as conn:
        rows = conn.execute(
            "SELECT * FROM pisma ORDER BY id DESC"
        ).fetchall()
    return [_row_to_pismo(r) for r in rows]


def search_pisma(query: str, db_path: Path = DB_PATH) -> list[Pismo]:
    """Wyszukuje pisma po sygnaturze, sądzie, typie, stronach i nazwie pliku."""
    like = f"%{query}%"
    with _connect(db_path) as conn:
        rows = conn.execute(
            """
            SELECT * FROM pisma
            WHERE sygnatura LIKE ? OR sad LIKE ? OR typ_pisma LIKE ?
               OR strona_powodowa LIKE ? OR strona_pozwana LIKE ? OR nazwa_pliku LIKE ?
            ORDER BY id DESC
            """,
            (like, like, like, like, like, like),
        ).fetchall()
    return [_row_to_pismo(r) for r in rows]


def delete_pismo(pismo_id: int, db_path: Path = DB_PATH) -> None:
    """Usuwa wpis z rejestru (NIE usuwa pliku z dysku)."""
    with _connect(db_path) as conn:
        conn.execute("DELETE FROM pisma WHERE id = ?", (pismo_id,))
        conn.commit()


def export_to_excel(path: str, db_path: Path = DB_PATH) -> None:
    """Eksportuje cały rejestr do pliku .xlsx pod wskazaną ścieżką."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    pisma = get_all_pisma(db_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Rejestr Kancelarii"

    headers = [EXCEL_HEADERS[c] for c in COLUMNS]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    kierunek_pl = {"IN": "przychodzące", "OUT": "wychodzące"}
    for p in pisma:
        row = []
        for col in COLUMNS:
            val = getattr(p, col)
            if col == "kierunek":
                val = kierunek_pl.get(val, val)
            row.append("" if val is None else val)
        ws.append(row)

    # Auto-szerokość kolumn (przybliżona).
    for idx, col in enumerate(COLUMNS, start=1):
        max_len = len(EXCEL_HEADERS[col])
        for p in pisma:
            val = getattr(p, col)
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = min(max_len + 2, 50)

    wb.save(path)
    logger.info("Wyeksportowano %d pism do %s", len(pisma), path)
